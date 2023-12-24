import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import CellRange
from openpyxl import Workbook, load_workbook

def read_validation_report(validation_file_path, sheet_name):
    if os.path.isfile(validation_file_path):
        return pd.read_excel(validation_file_path, sheet_name=sheet_name)
    else:
        return pd.DataFrame(columns=filtered_main_df2.columns)



def filter_and_append_data(validation_report, main_df):

    new_data_mask = ~main_df['Report Date'].isin(validation_report['Report Date'])
    
    if not main_df[new_data_mask].empty:
        existing_source_ids = validation_report[validation_report['Report Date'].isin(main_df['Report Date'])]['Source ID'].tolist()
        new_data = main_df[~((main_df['Source ID'].isin(existing_source_ids)) & (main_df['Report Date'].isin(validation_report['Report Date'])))].copy()
        
        if validation_report.empty:
            validation_report = new_data.copy()
        else:
            validation_report = pd.concat([validation_report, new_data], ignore_index=True)
    
    return validation_report
#Dropdown
def create_data_validation(sheet, column_range, values):
    formula = '"{}"'.format(",".join(values))
    data_validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    data_validation.add(column_range)
    sheet.add_data_validation(data_validation)


#Specific_paths
main_df = pd.read_excel(r"C:\Users\priya\Documents\SDL\04_main_db\main_db.xlsx")
validation_file_path = r"C:\Users\priya\Documents\SDL\03_report_to_validation\validation_report.xlsx"
lookup_df = pd.read_excel(r"C:\Users\priya\Documents\SDL\02_lookup\Source Nature Categorization_SPI NA Scheduling.xlsx", sheet_name='db')
# main_df = pd.read_excel(r"C:\Users\e600294\Xperi\Premkumar Jaganathan (CTR) - sdl_report\04_main_db\main_db.xlsx")
# lookup_df = pd.read_excel(r"C:\Users\e600294\Xperi\Premkumar Jaganathan (CTR) - sdl_report\02_lookup\Source Nature Categorization_SPI NA Scheduling.xlsx", sheet_name='db')
# validation_file_path = r"C:\Users\e600294\Xperi\Premkumar Jaganathan (CTR) - sdl_report\03_report_to_validation\validation_report.xlsx"
#dropping
main_df = main_df.drop(main_df[(main_df["Child Source?"] == "Y")].index)

main_df_channel = main_df.drop(main_df[(main_df['Logging Pattern / Source Nature'] == 'No Schedule/ Apply Source') |
                                       (main_df['Logging Pattern / Source Nature'] == 'Auto-Scheduling') |
                                       (main_df['Logging Pattern / Source Nature'] == 'Ignore') |
                                       (main_df['Logging Pattern / Source Nature'] == 'Ingest') |
                                       (main_df['Logging Pattern / Source Nature'] == 'Ingest Source') |
                                        (main_df['Logging Pattern / Source Nature'] == 'EMEA Source')|
                                        (main_df['Logging Pattern / Source Nature'] == 'Channel not found') |
                                         (main_df['Logging Pattern / Source Nature'] == 'Not Assigned to the reporter in Cosmo')].index)
main_df_channel = main_df_channel.drop(main_df_channel[(main_df_channel['Source Type'] =='Ignore')].index)
main_df_channel = main_df_channel.drop(main_df_channel[(main_df_channel['Source Type'] =='Ingest')].index)
#lookup_df = pd.read_excel(r"C:\Users\priya\Documents\SDL\02_lookup\Source Nature Categorization_SPI NA Scheduling.xlsx", sheet_name='db')
no_channel = lookup_df["Logging Pattern / Source Nature"]  ==   'Channel not found'
no_channel = lookup_df[no_channel]
main_df = main_df.drop(main_df[(main_df['Logging Pattern / Source Nature'] == 'No Schedule/ Apply Source')
                               | (main_df['Logging Pattern / Source Nature'] == 'Auto-Scheduling') |
                               (main_df['Logging Pattern / Source Nature'] == 'Ingest') |
                               (main_df['Logging Pattern / Source Nature'] == 'Ingest Source') |
                               (main_df['Logging Pattern / Source Nature'] == 'Channel not found') |
                               (main_df['Logging Pattern / Source Nature'] == 'Ignore') |
                                (main_df['Logging Pattern / Source Nature'] == 'EMEA Source')|
                                 (main_df['Logging Pattern / Source Nature'] == 'Not Assigned to the reporter in Cosmo')].index)
main_df = main_df.drop(main_df[(main_df['Source Type']   == 'Ignore')].index)
main_df = main_df.drop(main_df[(main_df['Source Type']   == 'Ingest')].index)
Filters_2 = (main_df["Lines\nTouched"] > 0) & (main_df["Lines\nTouched"] < 5) & (main_df["Days\nLogged"] > 7) & (main_df["Child Source?"] == "N")
filtered_main_df2 = main_df[Filters_2]                  
Filters = (main_df["Lines\nTouched"] == 0) 
Filters_1 = (main_df["Lines\nTouched"] >=100) & (main_df["Days\nLogged"] <= 7) 
filtered_main_df = main_df[Filters]
filtered_main_df1 = main_df[Filters_1]
filtered_main_df3 = main_df_channel[main_df_channel['Days\nLogged'] > 14] 
unique_dates = main_df["Report Date"].tolist()


required_columns = ['Report Date', 'Team', 'Reporter', 'Source ID', 'Source Name', 'SDL',
       'Child Source?', 'Logging Date', 'Day of Week', 'Lines\nTouched',
       'Days\nLogged', 'Line\nMulti-plier', 'Line\nWork\nUnits',
       'Day\nWork\nUnits', 'Total\nWork\nUnits',
       'Logging Pattern / Source Nature', 'Source Type', 'Reporter_Validation_status',
       'Reporter_ErrorCategory', 'Reporter_Sub_ErrorCategory','Reporter_Remarks',
       'Supervisor_Validation_status', 'Supervisor_ErrorCategory', 'Supervisor_Sub_ErrorCategory',
       'Supervisor_Remarks']

required_columns1 = ['Source ID', 'Source Name', 'Logging Pattern / Source Nature', 'Source Type','Remarks']

#read_validation_report
validation_report1 = read_validation_report(validation_file_path, sheet_name="Lines_Touched=0")
validation_report2 = read_validation_report(validation_file_path, sheet_name="Lines_Touched>100")
validation_report3 = read_validation_report(validation_file_path, sheet_name="Lines_Touched<5")
validation_report4 = read_validation_report(validation_file_path, sheet_name="Days_Logged>14")


# Filter and append new data
validation_report1 = filter_and_append_data(validation_report1, filtered_main_df)
validation_report2 = filter_and_append_data(validation_report2, filtered_main_df1)
validation_report3 = filter_and_append_data(validation_report3, filtered_main_df2)
validation_report4 = filter_and_append_data(validation_report4, filtered_main_df3)

# Remove duplicates based on the 'Source ID' and 'Report Date' columns
validation_report1.drop_duplicates(subset=['Source ID', 'Report Date'], keep='first', inplace=True)
validation_report2.drop_duplicates(subset=['Source ID', 'Report Date'], keep='first', inplace=True)
validation_report3.drop_duplicates(subset=['Source ID', 'Report Date'], keep='first', inplace=True)
validation_report4.drop_duplicates(subset=['Source ID', 'Report Date'], keep='first', inplace=True)

# Print a message if no new data is available
if validation_report1.equals(read_validation_report(validation_file_path, sheet_name="Lines_Touched=0")):
    print("No new data available to append to validation_report1")
if validation_report2.equals(read_validation_report(validation_file_path, sheet_name="Lines_Touched>100")):
    print("No new data available to append to validation_report2")
if validation_report3.equals(read_validation_report(validation_file_path, sheet_name="Lines_Touched<5")):
    print("No new data available to append to validation_report3")
if validation_report4.equals(read_validation_report(validation_file_path, sheet_name="Days_Logged>14")):
    print("No new data available to append to validation_report4")

# Save to Excel
with pd.ExcelWriter(validation_file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    validation_report1.to_excel(writer, columns=required_columns, sheet_name="Lines_Touched=0", index=False)
    validation_report2.to_excel(writer, columns=required_columns, sheet_name="Lines_Touched>100", index=False)
    validation_report3.to_excel(writer, columns=required_columns, sheet_name="Lines_Touched<5", index=False)
    validation_report4.to_excel(writer, columns=required_columns, sheet_name="Days_Logged>14", index=False)
#with pd.ExcelWriter(validation_file_path1) as writer:
    no_channel.to_excel(writer,  columns=required_columns1, sheet_name="channel_not_found", index=False)

workbook = openpyxl.load_workbook(validation_file_path)

# Create a list of valid validation status values
validation_status_list = ["", "Verified", "Channel already verified"]

# Create a list of valid error category values
error_category_list = ["", "Not_an_Error", "Technical_Error", "User_Error"]

# Create a list of sub_error_category values
sub_category_list = ["", "No_Changes, Template_updated", "Auto_Scheduling, Ingest", "Lines_touched", "Others"]

# Apply the data validation rule to the "Lines_Touched=0" sheet
ws_lines_touched_0 = workbook["Lines_Touched=0"]
validation_status_range_0 = f"R2:R{len(validation_report1)+1}"
remarks_range_0 = f"V2:V{len(validation_report1)+1}"
error_category_range_0 = f"S2:S{len(validation_report1)+1}"
error_category_range_00 = f"W2:W{len(validation_report1)+1}"
sub_category_range_0 = f"T2:T{len(validation_report1)+1}"
sub_category_range_1 = f"X2:X{len(validation_report1)+1}"
create_data_validation(ws_lines_touched_0, validation_status_range_0, validation_status_list)
create_data_validation(ws_lines_touched_0, remarks_range_0, validation_status_list)
create_data_validation(ws_lines_touched_0, error_category_range_0, error_category_list)
create_data_validation(ws_lines_touched_0, error_category_range_00, error_category_list)
create_data_validation(ws_lines_touched_0, sub_category_range_0, sub_category_list)
create_data_validation(ws_lines_touched_0, sub_category_range_1, sub_category_list)

# Apply the data validation rule to the "Lines_Touched>100" sheet
ws_lines_touched_1 = workbook["Lines_Touched>100"]
validation_status_range_1 = f"R2:R{len(validation_report2)+1}"
remarks_range_1 = f"V2:V{len(validation_report2)+1}"
error_category_range_01 = f"S2:S{len(validation_report2)+1}"
error_category_range_1 = f"W2:W{len(validation_report2)+1}"
sub_category_range_2 = f"T2:T{len(validation_report2)+1}"
sub_category_range_3 = f"X2:X{len(validation_report2)+1}"
create_data_validation(ws_lines_touched_1, validation_status_range_1, validation_status_list)
create_data_validation(ws_lines_touched_1, remarks_range_1, validation_status_list)
create_data_validation(ws_lines_touched_1, error_category_range_1, error_category_list)
create_data_validation(ws_lines_touched_1, error_category_range_01, error_category_list)
create_data_validation(ws_lines_touched_1, sub_category_range_2, sub_category_list)
create_data_validation(ws_lines_touched_1, sub_category_range_3, sub_category_list)

# Apply the data validation rule to the "Lines_Touched<5" sheet
ws_lines_touched_2 = workbook["Lines_Touched<5"]
validation_status_range_2 = f"R2:R{len(validation_report3)+1}"
remarks_range_2 = f"V2:V{len(validation_report3)+1}"
error_category_range_2 = f"S2:S{len(validation_report3)+1}"
error_category_range_02 = f"W2:W{len(validation_report3)+1}"
sub_category_range_4 = f"T2:T{len(validation_report3)+1}"
sub_category_range_5 = f"X2:X{len(validation_report3)+1}"
create_data_validation(ws_lines_touched_2, validation_status_range_2, validation_status_list)
create_data_validation(ws_lines_touched_2, remarks_range_2, validation_status_list)
create_data_validation(ws_lines_touched_2, error_category_range_2, error_category_list)
create_data_validation(ws_lines_touched_2, error_category_range_02, error_category_list)
create_data_validation(ws_lines_touched_2, sub_category_range_4, sub_category_list)
create_data_validation(ws_lines_touched_2, sub_category_range_5, sub_category_list)

# Apply the data validation rule to the "Days_Logged>14" sheet
ws_lines_touched_3 = workbook["Days_Logged>14"]
validation_status_range_3 = f"R2:R{len(validation_report4)+1}"
remarks_range_3 = f"V2:V{len(validation_report4)+1}"
error_category_range_3 = f"S2:S{len(validation_report4)+1}"
error_category_range_03 = f"W2:W{len(validation_report4)+1}"
sub_category_range_6 = f"T2:T{len(validation_report4)+1}"
sub_category_range_7 = f"X2:X{len(validation_report4)+1}"
create_data_validation(ws_lines_touched_3, validation_status_range_3, validation_status_list)
create_data_validation(ws_lines_touched_3, remarks_range_3, validation_status_list)
create_data_validation(ws_lines_touched_3, error_category_range_3, error_category_list)
create_data_validation(ws_lines_touched_3, error_category_range_03, error_category_list)
create_data_validation(ws_lines_touched_3, sub_category_range_6, sub_category_list)
create_data_validation(ws_lines_touched_3, sub_category_range_7, sub_category_list)

workbook.save(validation_file_path)

workbook = openpyxl.load_workbook(validation_file_path)
# Freeze the first row in all sheets (including 'channel_not_found' sheet)
for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    worksheet.freeze_panes = worksheet['A2']  # Freeze the first row
workbook.save(validation_file_path)
print("Done!")